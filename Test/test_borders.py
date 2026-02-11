"""Tests for xl_borders.borders module."""

import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from xl_borders import set_border
from xl_borders.borders import WEIGHT_STYLES


@pytest.fixture
def ws():
    """Create a fresh worksheet for each test."""
    wb = Workbook()
    return wb.active


class TestSetBorderSingleCell:
    def test_single_cell_all_sides_thin(self, ws):
        set_border(ws, "B2")

        cell = ws["B2"]
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"

    def test_single_cell_custom_style(self, ws):
        set_border(ws, "A1", style="thick")

        cell = ws["A1"]
        assert cell.border.left.style == "thick"
        assert cell.border.right.style == "thick"
        assert cell.border.top.style == "thick"
        assert cell.border.bottom.style == "thick"


class TestSetBorderRange:
    def test_range_outer_edges(self, ws):
        set_border(ws, "A1:C3")

        # Top-left corner
        assert ws["A1"].border.left.style == "thin"
        assert ws["A1"].border.top.style == "thin"

        # Top-right corner
        assert ws["C1"].border.right.style == "thin"
        assert ws["C1"].border.top.style == "thin"

        # Bottom-left corner
        assert ws["A3"].border.left.style == "thin"
        assert ws["A3"].border.bottom.style == "thin"

        # Bottom-right corner
        assert ws["C3"].border.right.style == "thin"
        assert ws["C3"].border.bottom.style == "thin"

    def test_range_inner_lines(self, ws):
        set_border(ws, "A1:C3")

        # Center cell should have inner borders
        center = ws["B2"]
        assert center.border.left.style == "thin"
        assert center.border.right.style == "thin"
        assert center.border.top.style == "thin"
        assert center.border.bottom.style == "thin"

    def test_outer_only_no_inner(self, ws):
        """Set outer borders but no inner lines."""
        no_side = Side(style=None)
        set_border(
            ws,
            "A1:C3",
            inner_horizontal=no_side,
            inner_vertical=no_side,
        )

        # Outer edges should be thin
        assert ws["A1"].border.left.style == "thin"
        assert ws["A1"].border.top.style == "thin"

        # Center cell should have no borders
        center = ws["B2"]
        assert center.border.left.style is None
        assert center.border.right.style is None
        assert center.border.top.style is None
        assert center.border.bottom.style is None


class TestSetBorderCustomSides:
    def test_mixed_styles(self, ws):
        thick = Side(style="thick")
        dashed = Side(style="dashed")

        set_border(ws, "A1:B2", left=thick, top=dashed)

        assert ws["A1"].border.left.style == "thick"
        assert ws["A1"].border.top.style == "dashed"
        # Right and bottom default to "thin"
        assert ws["B2"].border.right.style == "thin"
        assert ws["B2"].border.bottom.style == "thin"


class TestConvenienceParams:
    def test_outline_only(self, ws):
        """outline sets outer edges; inside stays at default style."""
        set_border(ws, "A1:C3", outline="medium")

        # Outer edges are medium
        assert ws["A1"].border.left.style == "medium"
        assert ws["A1"].border.top.style == "medium"
        assert ws["C3"].border.right.style == "medium"
        assert ws["C3"].border.bottom.style == "medium"

        # Inner lines fall back to default "thin"
        center = ws["B2"]
        assert center.border.left.style == "thin"
        assert center.border.top.style == "thin"

    def test_inside_only(self, ws):
        """inside sets inner lines; outer edges stay at default style."""
        set_border(ws, "A1:C3", inside="dashed")

        # Outer edges default to "thin"
        assert ws["A1"].border.left.style == "thin"
        assert ws["A1"].border.top.style == "thin"

        # Inner lines are dashed
        center = ws["B2"]
        assert center.border.left.style == "dashed"
        assert center.border.right.style == "dashed"
        assert center.border.top.style == "dashed"
        assert center.border.bottom.style == "dashed"

    def test_outline_and_inside(self, ws):
        """outline + inside covers all 6 positions."""
        set_border(ws, "A1:C3", outline="medium", inside="dotted")

        # Outer edges
        assert ws["A1"].border.left.style == "medium"
        assert ws["A1"].border.top.style == "medium"
        assert ws["C3"].border.right.style == "medium"
        assert ws["C3"].border.bottom.style == "medium"

        # Inner lines
        center = ws["B2"]
        assert center.border.left.style == "dotted"
        assert center.border.right.style == "dotted"
        assert center.border.top.style == "dotted"
        assert center.border.bottom.style == "dotted"

    def test_horizontal_and_vertical(self, ws):
        """horizontal/vertical split borders by orientation."""
        set_border(ws, "A1:C3", horizontal="thin", vertical="thick")

        # Top-left corner: left=thick (vertical), top=thin (horizontal)
        assert ws["A1"].border.left.style == "thick"
        assert ws["A1"].border.top.style == "thin"

        # Center cell: all lines from h/v
        center = ws["B2"]
        assert center.border.left.style == "thick"   # inner_vertical
        assert center.border.right.style == "thick"   # inner_vertical
        assert center.border.top.style == "thin"       # inner_horizontal
        assert center.border.bottom.style == "thin"    # inner_horizontal

    def test_individual_side_overrides_convenience(self, ws):
        """An explicit Side param takes priority over convenience params."""
        set_border(
            ws, "A1:C3",
            outline="medium",
            left=Side(style="double"),
        )

        # left overridden to double
        assert ws["A1"].border.left.style == "double"
        assert ws["A2"].border.left.style == "double"

        # other outer edges still medium
        assert ws["A1"].border.top.style == "medium"
        assert ws["C3"].border.right.style == "medium"
        assert ws["C3"].border.bottom.style == "medium"

    def test_single_cell_with_outline(self, ws):
        """On a single cell, outline applies to all 4 sides."""
        set_border(ws, "B2", outline="medium")

        cell = ws["B2"]
        assert cell.border.left.style == "medium"
        assert cell.border.right.style == "medium"
        assert cell.border.top.style == "medium"
        assert cell.border.bottom.style == "medium"


class TestCustomParam:
    def test_6_element_tuple(self, ws):
        """All 6 sides set via custom=(top, right, bottom, left, ih, iv)."""
        # top=thick, right=medium, bottom=thick, left=medium, ih=thin, iv=medium
        set_border(ws, "A1:C3", custom=(3, 2, 3, 2, 1, 2))

        # Top-left corner: top=thick, left=medium
        assert ws["A1"].border.top.style == "thick"
        assert ws["A1"].border.left.style == "medium"

        # Bottom-right corner: bottom=thick, right=medium
        assert ws["C3"].border.bottom.style == "thick"
        assert ws["C3"].border.right.style == "medium"

        # Center cell: inner borders (ih=thin, iv=medium)
        center = ws["B2"]
        assert center.border.top.style == "thin"
        assert center.border.bottom.style == "thin"
        assert center.border.left.style == "medium"
        assert center.border.right.style == "medium"

    def test_4_element_tuple(self, ws):
        """4-element tuple sets outer sides; inner defaults to no border."""
        # top=thick, right=medium, bottom=thick, left=medium
        set_border(ws, "A1:C3", custom=(3, 2, 3, 2))

        # Outer edges are set
        assert ws["A1"].border.top.style == "thick"
        assert ws["A1"].border.left.style == "medium"
        assert ws["C3"].border.bottom.style == "thick"
        assert ws["C3"].border.right.style == "medium"

        # Inner lines default to 0 (no border)
        center = ws["B2"]
        assert center.border.top.style is None
        assert center.border.bottom.style is None
        assert center.border.left.style is None
        assert center.border.right.style is None

    def test_custom_with_outline_override(self, ws):
        """outline (layer 3) overrides custom (layer 2) for outer edges."""
        set_border(ws, "A1:C3", custom=(1, 1, 1, 1, 1, 1), outline="thick")

        # Outer edges overridden by outline
        assert ws["A1"].border.top.style == "thick"
        assert ws["A1"].border.left.style == "thick"
        assert ws["C3"].border.bottom.style == "thick"
        assert ws["C3"].border.right.style == "thick"

        # Inner lines still from custom (thin)
        center = ws["B2"]
        assert center.border.top.style == "thin"
        assert center.border.left.style == "thin"

    def test_custom_with_individual_side_override(self, ws):
        """Individual Side param (layer 5) overrides custom (layer 2)."""
        set_border(
            ws, "A1:C3",
            custom=(1, 1, 1, 1, 1, 1),
            left=Side(style="double"),
        )

        # left overridden to double
        assert ws["A1"].border.left.style == "double"
        assert ws["A2"].border.left.style == "double"

        # Other outer edges still thin from custom
        assert ws["A1"].border.top.style == "thin"
        assert ws["C3"].border.right.style == "thin"

    def test_invalid_length_raises(self, ws):
        """custom tuple with length != 4 or 6 raises ValueError."""
        with pytest.raises(ValueError, match="4 or 6 elements"):
            set_border(ws, "A1:C3", custom=(1, 2, 3))

        with pytest.raises(ValueError, match="4 or 6 elements"):
            set_border(ws, "A1:C3", custom=(1, 2, 3, 1, 2))

    def test_invalid_weight_raises(self, ws):
        """custom tuple with weight not in WEIGHT_STYLES raises ValueError."""
        with pytest.raises(ValueError, match="not a valid weight"):
            set_border(ws, "A1:C3", custom=(1, 2, 5, 1))


class TestColorParam:
    def test_color_applies_to_all_sides(self, ws):
        """Base color is applied to every side."""
        set_border(ws, "A1:C3", color="FF0000")

        cell = ws["A1"]
        assert cell.border.left.color.rgb == "00FF0000"
        assert cell.border.top.color.rgb == "00FF0000"
        assert cell.border.left.style == "thin"

        center = ws["B2"]
        assert center.border.left.color.rgb == "00FF0000"
        assert center.border.top.color.rgb == "00FF0000"

    def test_color_with_style(self, ws):
        """color composes with style."""
        set_border(ws, "B2", style="thick", color="0000FF")

        cell = ws["B2"]
        assert cell.border.left.style == "thick"
        assert cell.border.left.color.rgb == "000000FF"

    def test_color_with_custom(self, ws):
        """color propagates through custom weights."""
        set_border(ws, "A1:C3", custom=(3, 2, 3, 2, 1, 1), color="00FF00")

        assert ws["A1"].border.top.style == "thick"
        assert ws["A1"].border.top.color.rgb == "0000FF00"
        assert ws["A1"].border.left.style == "medium"
        assert ws["A1"].border.left.color.rgb == "0000FF00"

        # Inner lines also get the color
        center = ws["B2"]
        assert center.border.top.style == "thin"
        assert center.border.top.color.rgb == "0000FF00"

    def test_color_with_outline(self, ws):
        """color propagates through outline."""
        set_border(ws, "A1:C3", outline="medium", color="FF0000")

        assert ws["A1"].border.left.style == "medium"
        assert ws["A1"].border.left.color.rgb == "00FF0000"

    def test_no_color_by_default(self, ws):
        """Without color param, sides have no color set."""
        set_border(ws, "B2")

        cell = ws["B2"]
        assert cell.border.left.color is None


class TestSideShorthand:
    def test_str_shorthand(self, ws):
        """A plain string sets the style, inheriting base color."""
        set_border(ws, "A1:C3", left="thick", color="FF0000")

        assert ws["A1"].border.left.style == "thick"
        assert ws["A1"].border.left.color.rgb == "00FF0000"
        # Other sides still default
        assert ws["A1"].border.top.style == "thin"

    def test_str_shorthand_no_base_color(self, ws):
        """A plain string without base color produces a Side with no color."""
        set_border(ws, "B2", left="thick")

        cell = ws["B2"]
        assert cell.border.left.style == "thick"
        assert cell.border.left.color is None

    def test_tuple_shorthand(self, ws):
        """A (style, color) tuple sets both, ignoring base color."""
        set_border(
            ws, "A1:C3",
            color="0000FF",
            left=("thick", "FF0000"),
        )

        # left uses tuple color, not base color
        assert ws["A1"].border.left.style == "thick"
        assert ws["A1"].border.left.color.rgb == "00FF0000"

        # Other sides use base color
        assert ws["A1"].border.top.color.rgb == "000000FF"

    def test_side_object_still_works(self, ws):
        """Passing a Side object directly still works as before."""
        set_border(ws, "B2", left=Side(style="double", color="00FF00"))

        cell = ws["B2"]
        assert cell.border.left.style == "double"
        assert cell.border.left.color.rgb == "0000FF00"

    def test_str_shorthand_overrides_outline(self, ws):
        """String shorthand at layer 5 overrides outline at layer 3."""
        set_border(ws, "A1:C3", outline="medium", left="double")

        assert ws["A1"].border.left.style == "double"
        assert ws["A1"].border.top.style == "medium"


class TestNumericRange:
    def test_single_cell_tuple(self, ws):
        """(row, col) targets a single cell."""
        set_border(ws, (2, 2))  # B2

        cell = ws.cell(row=2, column=2)
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"

    def test_range_tuple(self, ws):
        """((min_row, min_col), (max_row, max_col)) matches equivalent string."""
        set_border(ws, ((1, 1), (3, 3)))  # A1:C3

        # Same assertions as test_range_outer_edges
        assert ws["A1"].border.left.style == "thin"
        assert ws["A1"].border.top.style == "thin"
        assert ws["C1"].border.right.style == "thin"
        assert ws["C3"].border.bottom.style == "thin"

        # Inner
        center = ws["B2"]
        assert center.border.left.style == "thin"
        assert center.border.top.style == "thin"

    def test_range_tuple_with_kwargs(self, ws):
        """Numeric range works with all keyword params."""
        set_border(ws, ((1, 1), (3, 3)), outline="medium", inside="dashed")

        assert ws["A1"].border.left.style == "medium"
        assert ws["A1"].border.top.style == "medium"

        center = ws["B2"]
        assert center.border.left.style == "dashed"
        assert center.border.top.style == "dashed"

    def test_numeric_matches_string(self, ws):
        """Numeric range produces identical result to equivalent string range."""
        wb2 = Workbook()
        ws2 = wb2.active

        set_border(ws, "A1:C3", style="thick")
        set_border(ws2, ((1, 1), (3, 3)), style="thick")

        for row in range(1, 4):
            for col in range(1, 4):
                b1 = ws.cell(row=row, column=col).border
                b2 = ws2.cell(row=row, column=col).border
                assert b1.left.style == b2.left.style
                assert b1.right.style == b2.right.style
                assert b1.top.style == b2.top.style
                assert b1.bottom.style == b2.bottom.style

    def test_invalid_type_raises(self, ws):
        """Invalid cell_range type raises TypeError."""
        with pytest.raises(TypeError, match="cell_range must be"):
            set_border(ws, 42)  # type: ignore[arg-type]


class TestPreserveCellFormatting:
    """set_border() must not overwrite font, fill, or alignment."""

    def test_preserves_font(self, ws):
        ws["B2"].font = Font(bold=True, color="FF0000", size=14)
        set_border(ws, "A1:C3")

        font = ws["B2"].font
        assert font.bold is True
        assert font.color.rgb == "00FF0000"
        assert font.size == 14

    def test_preserves_fill(self, ws):
        ws["B2"].fill = PatternFill(fgColor="FFFF00", patternType="solid")
        set_border(ws, "A1:C3")

        fill = ws["B2"].fill
        assert fill.fgColor.rgb == "00FFFF00"
        assert fill.patternType == "solid"


class TestPreserveBorderColor:
    """Border color merging: preserve existing color when not specified."""

    def test_preserves_color_when_no_color_param(self, ws):
        """Existing red border -> set_border() without color -> red preserved."""
        ws["B2"].border = Border(
            left=Side(style="thin", color="FF0000"),
            top=Side(style="thin", color="FF0000"),
        )
        set_border(ws, "B2")

        cell = ws["B2"]
        assert cell.border.left.style == "thin"
        assert cell.border.left.color.rgb == "00FF0000"
        assert cell.border.top.color.rgb == "00FF0000"

    def test_overwrites_color_when_color_param_set(self, ws):
        """Existing red border -> set_border(color='0000FF') -> blue."""
        ws["B2"].border = Border(
            left=Side(style="thin", color="FF0000"),
        )
        set_border(ws, "B2", color="0000FF")

        cell = ws["B2"]
        assert cell.border.left.style == "thin"
        assert cell.border.left.color.rgb == "000000FF"

    def test_preserves_color_with_style_change(self, ws):
        """Change style but keep existing color when color param not set."""
        ws["B2"].border = Border(
            left=Side(style="thin", color="FF0000"),
        )
        set_border(ws, "B2", style="thick")

        cell = ws["B2"]
        assert cell.border.left.style == "thick"
        assert cell.border.left.color.rgb == "00FF0000"

    def test_tuple_shorthand_color_overrides_existing(self, ws):
        """Explicit color in (style, color) tuple overrides existing."""
        ws["B2"].border = Border(
            left=Side(style="thin", color="FF0000"),
        )
        set_border(ws, "B2", left=("thick", "00FF00"))

        cell = ws["B2"]
        assert cell.border.left.style == "thick"
        assert cell.border.left.color.rgb == "0000FF00"


class TestPreserveExistingBorders:
    """Existing borders are kept when set_border() doesn't configure that side."""

    def test_custom_4_preserves_unset_inner_borders(self, ws):
        """4-element custom sets outer only; existing inner borders preserved."""
        # Pre-set inner borders on center cell
        ws["B2"].border = Border(
            left=Side(style="dashed", color="FF0000"),
            top=Side(style="dashed", color="FF0000"),
        )
        set_border(ws, "A1:C3", custom=(3, 3, 3, 3))

        # Outer edges set to thick
        assert ws["A1"].border.left.style == "thick"
        assert ws["A1"].border.top.style == "thick"

        # Center cell: inner sides resolve to style=None -> existing preserved
        center = ws["B2"]
        assert center.border.left.style == "dashed"
        assert center.border.left.color.rgb == "00FF0000"
        assert center.border.top.style == "dashed"

    def test_outline_only_preserves_existing_inner(self, ws):
        """outline + no inside: inner lines keep default style, not cleared."""
        ws["B2"].border = Border(
            left=Side(style="double"),
            top=Side(style="double"),
        )
        set_border(ws, "A1:C3", outline="thick")

        # Outer edges are thick
        assert ws["A1"].border.left.style == "thick"

        # Inner: style defaults to "thin" (from base style param), overrides existing
        center = ws["B2"]
        assert center.border.left.style == "thin"
        assert center.border.top.style == "thin"

    def test_none_side_preserves_existing(self, ws):
        """Explicitly passing Side(style=None) preserves existing border."""
        ws["B2"].border = Border(left=Side(style="thick", color="FF0000"))
        set_border(ws, "B2", left=Side(style=None))

        cell = ws["B2"]
        assert cell.border.left.style == "thick"
        assert cell.border.left.color.rgb == "00FF0000"
